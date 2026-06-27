import pandas as pd
import numpy as np

# Load Jun 23 sheet exactly up to row 98 (header is at 1, so rows 2 to 98 is 97 rows total)
df_jun23 = pd.read_excel('R037_External RGOB debt_Sector mapping.xlsx', 
                         sheet_name='Jun 23', 
                         header=1,
                         nrows=97)  # rows 2-98 inclusive is 97 rows

# Drop completely empty rows
df_jun23 = df_jun23.dropna(how='all')

# Forward fill Donor/Lender
df_jun23['Donor/Lender'] = df_jun23['Donor/Lender'].ffill()

# Keep only rows where Project Name is not NaN
df_jun23 = df_jun23.dropna(subset=['Project Name'])

# Remove duplicates
df_jun23 = df_jun23.drop_duplicates()

# Define the columns we need to unpivot (the value columns)
value_columns = [
    'Loan Committed',
    'Loan Disbursed', 
    'Principal Repayment made', 
    'Interest Paid',
    'Outstanding Loan Balance for Repayment'
]

# Unpivot (melt) the data to long format
df_melted = df_jun23.melt(
    id_vars=['Donor/Lender', 'Project Name', 'Sector ', 'Currency'],
    value_vars=value_columns,
    var_name='loan_disbursements',
    value_name='value'
)

# Ensure value is numeric
df_melted['value'] = pd.to_numeric(df_melted['value'], errors='coerce').fillna(0)

# Now let's create the reference file columns
# Let's get some static values from the reference file
df_ref = pd.read_excel('Copy of R037_Datasheet-2021.xlsx', sheet_name='Sheet1')

# Create the final DataFrame
df_final = pd.DataFrame()

# Copy static fields (we'll use the first row's values for most)
df_final['file_nm'] = ['R037_Ex_RGOB_debt_Jun23.json'] * len(df_melted)
df_final['dataidentifier'] = range(1, len(df_melted) + 1)
df_final['return_id'] = 'R037'
df_final['entity_id'] = df_ref['entity_id'].iloc[0]

# Convert dates to datetime64[ns] (same as reference file)
report_start_date = pd.to_datetime('2023-06-01')
report_end_date = pd.to_datetime('2023-06-30')
df_final['report_start_dt'] = [report_start_date] * len(df_melted)
df_final['daily_report_end_dt'] = [report_end_date] * len(df_melted)
df_final['monthly_report_end_dt'] = [report_end_date] * len(df_melted)
df_final['pe_id'] = df_ref['pe_id'].iloc[0]
df_final['mis_date'] = [report_end_date] * len(df_melted) 
df_final['report_date'] = [report_end_date] * len(df_melted)

# Convert columns to datetime64[ns]
for col in ['report_start_dt', 'daily_report_end_dt', 'monthly_report_end_dt', 'mis_date', 'report_date']:
    df_final[col] = pd.to_datetime(df_final[col])

# Now map the fields from our melted data
df_final['donor_lender'] = df_melted['Donor/Lender']
df_final['project_name'] = df_melted['Project Name']
df_final['sector'] = df_melted['Sector ']
df_final['loan_disbursements'] = df_melted['loan_disbursements']
df_final['currency'] = df_melted['Currency']
df_final['value'] = df_melted['value']

# Map de_id and dm_id based on loan_disbursements
de_id_map = {
    'Loan Committed': 'rma_DE000061',
    'Loan Disbursed': 'rma_DE000061',
    'Principal Repayment made': 'rma_DE000061',
    'Interest Paid': 'rma_DE000061',
    'Outstanding Loan Balance for Repayment': 'rma_DE000061'
}
dm_id_map = {
    'Loan Committed': 'rma_DM000347',
    'Loan Disbursed': 'rma_DM000348',
    'Principal Repayment made': 'rma_DM000349',
    'Interest Paid': 'rma_DM000350',
    'Outstanding Loan Balance for Repayment': 'rma_DM000351'
}
df_final['de_id'] = df_melted['loan_disbursements'].map(de_id_map).fillna('rma_DE000061')
df_final['dm_id'] = df_melted['loan_disbursements'].map(dm_id_map).fillna('rma_DM000347')
df_final['currency_axis'] = 'Convertible currencies'

# Reorder columns to match the reference file exactly
ref_columns = df_ref.columns.tolist()
df_final = df_final[ref_columns]

# Remove duplicates from final DataFrame
df_final = df_final.drop_duplicates()

print("=== Final DataFrame ===")
print(df_final.head(20).to_string())

# Save to Excel using openpyxl to match reference file format
from openpyxl import load_workbook
from openpyxl.styles import numbers

# First save with pandas to get the data in
df_final.to_excel('junclean data.xlsx', index=False, sheet_name='Sheet1')

# Now open with openpyxl to apply the number format
wb = load_workbook('junclean data.xlsx')
ws = wb.active

# Get column letters for date columns
ref_columns = df_ref.columns.tolist()
date_cols = ['report_start_dt', 'daily_report_end_dt', 'monthly_report_end_dt', 'mis_date', 'report_date']

# Function to convert column index to Excel column letter (A, B, ..., Z, AA, AB, ...)
def col_idx_to_letter(n):
    string = ''
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

for col_name in date_cols:
    col_idx = ref_columns.index(col_name)
    col_letter = col_idx_to_letter(col_idx + 1)  # +1 because Excel is 1-indexed
    for cell in ws[col_letter]:
        cell.number_format = 'yyyy-mm-dd'  # User wants this format

wb.save('junclean data.xlsx')

print("\nCleaned data saved as junclean data.xlsx")
